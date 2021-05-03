from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws.title='Sample'
ws['A1'].value='Hello text'
for i in range(2,10):
    ws.cell(row=i,column=1).value='No'+str(i-1)
wb.save('data.xlsx')
print('saved')
